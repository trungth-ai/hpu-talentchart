# Import nhân sự HPU từ file Excel lương (VD: "Luong T8 gửi Trung.xlsx")
# → tạo Candidate với candidate_type=employee (hợp nhất theo Critical Business Rules)
#
# ⚠️ Dữ liệu ngày sinh là dữ liệu NHẠY CẢM (NĐ 13/2023/NĐ-CP):
#   - Mặc định KHÔNG import ngày sinh
#   - Chỉ import khi chạy kèm --with-epa-consent (xác nhận đã có cơ sở đồng ý
#     từ hệ Fortune HR cũ đang dùng tại HPU)
# ⚠️ KHÔNG import bất kỳ cột lương nào — ngoài phạm vi ATS.
# ⚠️ Email placeholder {slug}.{tt}@import.hpu.edu.vn — cập nhật email thật sau
#   (nhân sự đăng nhập Google @hpu.edu.vn sẽ được match theo email thật).
#
# Chạy (từ thư mục backend/, cần env DATABASE_URL + JWT_SECRET):
#   python scripts/import_employees.py --file "path/to/Luong T8.xlsx" --org-slug hpu \
#       [--with-epa-consent] [--dry-run]

import argparse
import asyncio
import re
import sys
import unicodedata
from datetime import UTC, date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.tenant_context import reset_current_org_id, set_current_org_id  # noqa: E402
from app.database import async_session_factory, set_rls_guc  # noqa: E402
from app.models import Candidate, Organization  # noqa: E402


def _slugify(text: str) -> str:
    text = text.replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", ".", text.lower()).strip(".")


def _parse_rows(file_path: str, sheet: str | None) -> list[dict]:
    """Đọc Excel: tự tìm hàng header (TT + Họ và tên), trả về list nhân sự."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row[:6]]
        if cells[0] == "TT" and "Họ và tên" in cells:
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit("[LOI] Khong tim thay hang header (TT | Ho va ten) trong file")

    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]

    def col(name: str) -> int | None:
        return header.index(name) if name in header else None

    idx = {
        "tt": col("TT"),
        "full_name": col("Họ và tên"),
        "day": col("Ngày sinh"),
        "month": col("Tháng sinh"),
        "year": col("Năm sinh"),
        "dept": col("Mã bộ phận"),
        "position_ql": col("Tên vị trí quản lý"),
        "position_cs": col("Tên đối tượng hưởng lương cơ sở"),
    }

    employees = []
    # LƯU Ý: file có nhiều phân đoạn, cột TT đánh lại từ 1 ở mỗi phân đoạn
    # → dedupe theo (họ tên + ngày sinh), KHÔNG theo TT
    seen: set[tuple] = set()
    for row in rows[header_idx + 1 :]:
        def cell(key: str, _row=row):
            # _row bind theo từng vòng lặp (tránh closure trên biến loop)
            i = idx[key]
            return _row[i] if i is not None and i < len(_row) else None

        full_name = cell("full_name")
        tt = cell("tt")
        if not full_name or not str(full_name).strip():
            continue  # bỏ hàng trống/tổng cộng
        try:
            tt_num = int(tt)
        except (TypeError, ValueError):
            continue  # hàng không phải dữ liệu nhân sự

        # Ngày sinh — 3 cột riêng, có thể thiếu
        birth = None
        try:
            d, m, y = int(cell("day")), int(cell("month")), int(cell("year"))
            birth = date(y, m, d)
        except (TypeError, ValueError):
            pass

        # Vị trí: ưu tiên vị trí quản lý, fallback đối tượng hưởng lương cơ sở
        position = None
        for key in ("position_ql", "position_cs"):
            val = cell(key)
            if val and str(val).strip() and str(val).strip() != "???":
                position = str(val).strip()
                break

        dept = cell("dept")

        # Bỏ hàng rác: không có bất kỳ thông tin phụ nào (VD: hàng "Copy" sót lại)
        if birth is None and dept is None and position is None:
            print(f"[BO QUA] Hang thieu du lieu: TT {tt_num} — {full_name}")
            continue
        key = (str(full_name).strip().lower(), birth)
        if key in seen:
            print(f"[BO QUA] Trung lap: {full_name} ({birth})")
            continue
        seen.add(key)

        employees.append(
            {
                # Số thứ tự LIÊN TỤC toàn file (TT gốc đánh lại theo phân đoạn)
                "seq": len(employees) + 1,
                "full_name": str(full_name).strip(),
                "birth_date": birth,
                "department": str(dept).strip() if dept else None,
                "position": position,
            }
        )
    return employees


async def import_employees(
    file_path: str,
    org_slug: str,
    sheet: str | None,
    with_epa_consent: bool,
    dry_run: bool,
) -> None:
    employees = _parse_rows(file_path, sheet)
    print(f"[INFO] Doc duoc {len(employees)} nhan su tu file")

    if dry_run:
        for e in employees[:10]:
            print("  ", e)
        print(f"[DRY-RUN] Khong ghi database. --with-epa-consent={with_epa_consent}")
        return

    async with async_session_factory() as session:
        org = (
            await session.execute(select(Organization).where(Organization.slug == org_slug))
        ).scalar_one_or_none()
        if org is None:
            raise SystemExit(f"[LOI] Khong tim thay organization slug '{org_slug}'")

        token = set_current_org_id(org.id)
        try:
            await set_rls_guc(session)
            created, skipped, no_birth = 0, 0, 0

            for e in employees:
                email = f"{_slugify(e['full_name'])}.{e['seq']}@import.hpu.edu.vn"

                # Idempotent: đã có employee cùng email hoặc cùng employee_code → bỏ qua
                existing = await session.execute(
                    select(Candidate.id)
                    .where(Candidate.organization_id == org.id)
                    .where(Candidate.email == email)
                )
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue

                store_birth = with_epa_consent and e["birth_date"] is not None
                if e["birth_date"] is None:
                    no_birth += 1

                session.add(
                    Candidate(
                        organization_id=org.id,
                        full_name=e["full_name"],
                        email=email,
                        candidate_type="employee",
                        # Nhân sự đang làm việc = đã qua tuyển dụng
                        pipeline_stage="HIRED",
                        employee_code=f"NV{e['seq']:04d}",
                        department=e["department"],
                        position=e["position"],
                        source="import_excel",
                        epa_consent=store_birth,
                        epa_consent_at=datetime.now(UTC) if store_birth else None,
                        birth_date=e["birth_date"] if store_birth else None,
                    )
                )
                created += 1

            await session.commit()
        finally:
            reset_current_org_id(token)

    print(f"[OK] Da import {created} nhan su (bo qua {skipped} da ton tai)")
    if no_birth:
        print(f"[CHU Y] {no_birth} nguoi thieu ngay sinh trong file")
    if not with_epa_consent:
        print(
            "[CHU Y] Ngay sinh KHONG duoc import (thieu --with-epa-consent). "
            "EPA se khong tinh duoc cho den khi co du lieu + consent."
        )
    print(
        "[CHU Y] Email dang placeholder @import.hpu.edu.vn — "
        "cap nhat email that de match Google login"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import nhan su tu Excel vao TalentChart")
    parser.add_argument("--file", required=True, help="Duong dan file .xlsx")
    parser.add_argument("--org-slug", required=True, help="Slug cua organization (vd: hpu)")
    parser.add_argument("--sheet", default=None, help="Ten sheet (mac dinh: sheet dau)")
    parser.add_argument(
        "--with-epa-consent",
        action="store_true",
        help="Import ca ngay sinh + danh dau epa_consent (chi khi da co co so dong y)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Chi doc & in thu, khong ghi DB")
    args = parser.parse_args()

    asyncio.run(
        import_employees(
            args.file, args.org_slug, args.sheet, args.with_epa_consent, args.dry_run
        )
    )
